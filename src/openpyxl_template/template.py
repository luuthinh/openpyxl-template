# -*- coding:utf-8 -*-
"""
    @author: luuthinh2705@gmail.com
"""

from os import PathLike
from typing import Any, Optional, IO, Union, Dict, Set
from copy import copy, deepcopy
from itertools import chain, groupby

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter
from openpyxl.drawing.image import Image

from jinja2 import Environment, Template, meta
from jinja2.exceptions import TemplateError

from .utils import (
    check_jinja_syntax,
    render,
    remove_loop_syntax,
    get_loop_block,
    cleaning,
    has_parent
)

safe_globals = {"__builtins__": {}}


class ExcelTemplate(object):
    """ Class for managing excel file as they were jinja2 templates"""

    def __init__(self, template_file: Union[IO[bytes], str, PathLike], max_row_tmpl, max_col_tmpl) -> None:
        self.template_file = template_file
        self.workbook = None
        self.is_rendered = False
        self.is_saved = False
        self.max_row_tmpl = max_row_tmpl
        self.max_col_tmpl = max_col_tmpl

    def init_workbook(self, reload: bool=True):
        if not self.workbook or (self.is_rendered and reload):
            self.workbook = load_workbook(self.template_file)
            self.is_rendered = False

    def render_init(self):
        self.init_workbook()
    
    def __getattr__(self, name):
        return getattr(self.workbook, name)
    
    def get_workbook(self):
        self.init_workbook()
        return self.workbook
    
    def render(self, 
        context: Dict[str, Any], 
    ) -> None:
        self.render_init()
        for sheet in self.workbook.sheetnames:
            ws_source = self.workbook[sheet]
            blocks, context_values = self.find_block(ws_source)
            blocks.sort(key=lambda b: b["min_row"])

            # Update the context with additional values from context_values
            for ctx in context_values:
                try:
                    context[ctx["key"]] = eval(cleaning(ctx["attr"]), safe_globals, context)
                except Exception as e:
                    raise e
                
            ws_des = self.workbook.copy_worksheet(ws_source)
            for img in ws_source._images:
                ws_des.add_image(deepcopy(img))
            ws_des.title = ws_source.title + "- final"
            # unmerge_cell
            self.unmerge_cells(ws_des, blocks)
            max_row_by_rows = 0
            above_blocks = []                

            for key, block_same_row in groupby(blocks, lambda b: b["min_row"]):
                list_block_same_row = sorted(block_same_row, key=lambda b: b["min_col"])
                offset_col = 0
                max_row_by_rows = 0
                for left_seq, block in enumerate(list_block_same_row):
                    block["top_blocks"] = list(filter(lambda b: b["min_col"] <= block["min_col"] <= b["max_col"] or b["min_col"] <= block["max_col"] <= b["max_col"], above_blocks))

                    offset_row = self.get_offset_row(block)
                    if left_seq == 0:
                        offset_col = list_block_same_row[left_seq].get(
                            "min_col", 0)
                    (
                        offset_col,
                        max_row_by_block,
                        increase_row_by_block,
                    ) = self._render_block(
                        ws_source,
                        ws_des,
                        block,
                        context,
                        offset_col,
                        offset_row,
                        max_row_by_rows,
                        wb=self.workbook,
                        left_block=None if left_seq == 0 else list_block_same_row[left_seq - 1],
                    )
                    block["increace_row"] = increase_row_by_block

                    if max_row_by_block > max_row_by_rows:
                        max_row_by_rows = max_row_by_block

                above_blocks = list_block_same_row
            # merge_cell
            self.merge_cell(ws_des)

        # for sheet in self.workbook.sheetnames:
        #     self.workbook.remove(self.workbook[f"{sheet}"])

    def save(self, filename: Union[IO[bytes], str, PathLike]) -> None:
        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        print(self.workbook.sheetnames)
        self.workbook.save(filename)
        self.is_saved = True              

    ######################## Preprocessing block ##########################
    def find_block(self, ws_template):
        blocks = []
        context_values = []
        all_cell = chain.from_iterable(
            ws_template.iter_rows(
                min_row=1,
                max_row=self.max_row_tmpl,
                min_col=1,
                max_col=self.max_col_tmpl,
            )
        )
        for cell in filter(lambda c: c.value, all_cell):
            mergedCell = next(
                filter(
                    lambda mCell: cell.coordinate in mCell,
                    ws_template.merged_cells.ranges,
                ),
                None,
            )

            block_for_splits = get_loop_block(cell.value)
            if len(block_for_splits) == 0:
                continue
            try:
                for block_for_split in block_for_splits:
                    if block_for_split[0] == "setValue":
                        if len(block_for_split) == 3:
                            context_values.append({
                                "attr": block_for_split[1],
                                "key": block_for_split[2]
                            })
                    
                    if block_for_split[0] == "table":
                        blocks.append(
                            {
                                "min_row": mergedCell.min_row if mergedCell else cell.row,
                                "min_col": mergedCell.min_col if mergedCell else cell.column,
                                "max_row": mergedCell.max_row if mergedCell else cell.row,
                                "max_col": mergedCell.max_col if mergedCell else cell.column,
                                "type": block_for_split[0],
                                "code": block_for_split[1],
                                "attr": block_for_split[2] if len(block_for_split) >= 3 else None,
                                "child": None,
                                "nested_blocks": [],
                                "closed": True
                            }
                        )

                    if block_for_split[0] in ["forCol", "forRow", "static"]:
                        blocks.append(
                            {
                                "min_row": mergedCell.min_row if mergedCell else cell.row,
                                "min_col": mergedCell.min_col if mergedCell else cell.column,
                                "max_row": mergedCell.max_row if mergedCell else cell.row,
                                "max_col": mergedCell.max_col if mergedCell else cell.column,
                                "type": block_for_split[0],
                                "code": block_for_split[1],
                                "attr": block_for_split[2] if len(block_for_split) >= 3 else None,
                                "child": block_for_split[3] if len(block_for_split) == 4 else None,
                                "nested_blocks": []
                            }
                        )

                    if block_for_split[0] in ["endCol", "endRow", "endStatic"]:
                        ex_result = next(
                            filter(lambda x: x["code"] ==
                                   block_for_split[1], blocks), None
                        )
                        if ex_result:
                            ex_result.update(
                                {
                                    "max_row": mergedCell.max_row if mergedCell else cell.row,
                                    "max_col": mergedCell.max_col if mergedCell else cell.column,
                                    "closed": True
                                }
                            )
                    cell.value = remove_loop_syntax(cell.value)
            except:
                continue
        msg = "Error:"
        flag_miss_closing = False
        for block in blocks:
            if not block.get("closed", False):
                flag_miss_closing = True
                msg += f"\n {block['code']} - The block is missing a closing tag."
            if block["type"] == "forRow" and block["min_row"] != block["max_row"]:
                msg += f"\n - {block['code']} - The block must be same row."
        if flag_miss_closing:
            raise Exception(msg)
        
        return self._process_block(blocks), context_values
    
    def _process_block(self, blocks, ex_blocks=[]):
        final_blocks = deepcopy(ex_blocks)
        for child in blocks:
            flag = True
            for parent in blocks:
                if parent != child and has_parent(parent, child):
                    flag = False
                    continue
            if flag:
                for i, block in enumerate(blocks):
                    if block["code"] == parent["code"]:
                        final_blocks.append(blocks.pop(i))
                        break
        final_blocks = self._nested_block(final_blocks)
        if not len(blocks):
            return final_blocks
        else:
            return self._process_block(blocks, final_blocks)

    def _remove_duplicate_nested_items(self, blocks):
        nested_codes = {
            nested_block["code"]
            for block in blocks if "nested_blocks" in block
            for nested_block in block["nested_blocks"]
        }
        return [block for block in blocks if block["code"] not in nested_codes]

    def _nested_block(self, blocks):     
        for parent in blocks:
            if parent["type"] not in ["forCol", "static"]:
                continue
            for child in blocks:
                if child["type"] != "forRow":
                    continue
                if has_parent(parent, child):
                    parent["nested_blocks"].append(child)

        return self._remove_duplicate_nested_items(blocks)
    
    def unmerge_cells(self, ws_des, blocks):
        ranges = ws_des.merged_cells.ranges
        for block in blocks:
            min_row = block.get("min_row", 0)
            min_col = block.get("min_col", 0)

            max_row = block.get("max_row", 0)
            max_col = block.get("max_col", 0)

            for row in ws_des.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    cell.value = None
                    cell.style = "Normal"
                    ws_des.row_dimensions[cell.row].height = 15
                    mergedCell = next(
                        filter(
                            lambda mCell: cell.coordinate in mCell,
                            ranges,
                        ),
                        None,
                    )
                    if not mergedCell:
                        continue

                    ws_des.unmerge_cells(mergedCell.coord)

    ######################## Common Function ##########################
    def get_offset_row(self, block):
        total = block.get("increace_row", 0)
        top_blocks = block.get("top_blocks", [])
        if top_blocks:
            block_with_max_row = max(top_blocks, key=lambda b: b['increace_row'])
            total += self.get_offset_row(block_with_max_row)
        return total
    
    def _find_next_row(self, cur_rows, merge_rows, result):
        list_col = [a[1] for a in cur_rows]
        next_rows = next(
            filter(
                lambda rows: rows[0][0] == cur_rows[-1][0] + 1
                and rows[0][2] == cur_rows[0][2]
                and set([row[1] for row in rows]).issuperset(set(list_col)),
                merge_rows,
            ),
            None,
        )
        if next_rows:
            merge_rows.pop(merge_rows.index(next_rows))
            try:
                result.pop(result.index(cur_rows))
            except:
                pass
            result.append(cur_rows + next_rows)
            self._find_next_row(result[-1], merge_rows, result)
        else:
            try:
                result.pop(result.index(cur_rows))
            except:
                pass
            result.append(cur_rows)    

    def _cleaning_wb(self):
        before_sheets = []
        after_sheets = []
        for ws_source in self.workbook.worksheets:
            before_sheets.append(ws_source.title)
            after_sheets.append(ws_source.title + "-copy")
            ws_des = self.workbook.create_sheet(ws_source.title + "-copy")
            for r, row in enumerate(
                ws_source.iter_rows(
                    min_row=0, min_col=0, max_row=self.max_row_tmpl, max_col=self.max_col_tmpl
                )
            ):
                for c, cell in enumerate(row):
                    new_cell = ws_des.cell(row=r + 1, column=c + 1)
                    new_cell.value = cell.value
                    new_cell.data_type = cell.data_type

                    if cell.has_style:
                        new_cell._style = copy(cell._style)

                    if cell.hyperlink:
                        new_cell._hyperlink = copy(cell.hyperlink)

                    if cell.comment:
                        new_cell.comment = copy(cell.comment)
            self._copy_dimensions(ws_source, ws_des)
            ws_des.sheet_format = copy(ws_source.sheet_format)
            ws_des.sheet_properties = copy(ws_source.sheet_properties)
            ws_des.merged_cells = copy(ws_source.merged_cells)
            ws_des.page_margins = copy(ws_source.page_margins)
            ws_des.page_setup = copy(ws_source.page_setup)
            ws_des.print_options = copy(ws_source.print_options)
            for img in ws_source._images:
                ws_des.add_image(deepcopy(img))

        for sheet in before_sheets:
            self.workbook.remove(self.workbook[f"{sheet}"])
        for sheet in after_sheets:
            self.workbook[sheet].title = self.workbook[sheet].title[:-5]

    def _copy_dimensions(self, source, target):
        for attr in ("row_dimensions", "column_dimensions"):
            src = getattr(source, attr)
            tar = getattr(target, attr)
            for key, dim in src.items():
                tar[key] = copy(dim)
                tar[key].worksheet = target

    def _set_dimensions(self, ws, cell, height=None, width=None):
        if ws.row_dimensions[cell.row].height is None or ws.row_dimensions[cell.row].height != (height or 0):
            ws.row_dimensions[cell.row].height = height
        if ws.column_dimensions[cell.column_letter].width is None or ws.column_dimensions[cell.column_letter].width < (width or 0):
            ws.column_dimensions[cell.column_letter].width = width                

    def merge_cell(self, ws):
        cells = []
        merge_rows = []
        results = []
        if ws.max_column < 2 or ws.max_row < 2:
            return

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                value = ws[f"{get_column_letter(c)}{r}"].value
                if isinstance(value, str):
                    if "merge_cell" in value:
                        cells.append((r, c, "merge_cell"))
                    if "merge_row" in value:
                        cells.append((r, c, "merge_row"))
                    if "merge_col" in value:
                        cells.append((r, c, "merge_col"))

        if not cells:
            return

        # Merge cell in row
        for r, g_cells in groupby(cells, lambda m: m[0]):
            l_cell = list(g_cells)
            merges = []
            prev_cell = 0
            for cur_cell in l_cell:
                if not len(merges) or (
                    cur_cell[1] -
                        prev_cell[1] == 1 and cur_cell[2] == cur_cell[2]
                ):
                    if cur_cell[2] == "merge_col":
                        merge_rows.append([cur_cell])
                    else:
                        merges.append(cur_cell)
                else:
                    merge_rows.append(merges)
                    merges = [cur_cell]

                if len(merges) and cur_cell[1] == l_cell[-1][1]:
                    merge_rows.append(merges)
                prev_cell = cur_cell

        # Merge cell in column
        for merge_row in merge_rows:
            if merge_row[0][2] == "merge_row":
                results.append(merge_row)
                continue
            self.find_next_row(merge_row, merge_rows, results)

        for result in results:
            list_row = [r[0] for r in result]
            list_col = [c[1] for c in result]
            min_row = min(list_row)
            max_row = max(list_row)
            min_col = min(list_col)
            max_col = max(list_col)
            first_cell = ws.cell(row=min_row, column=min_col)

            if min_row == max_row and min_col == max_col:
                if "merge_col" in first_cell.value and min_row > 1:
                    ws.merge_cells(
                        start_row=min_row - 1,
                        start_column=min_col,
                        end_row=max_row,
                        end_column=max_col,
                    )
                    continue
                if "merge_cell" in first_cell.value and min_col > 1:
                    ws.merge_cells(
                        start_row=min_row,
                        start_column=min_col - 1,
                        end_row=max_row,
                        end_column=max_col,
                    )
                    continue

            if min_row != max_row and min_col != max_col:
                ws.merge_cells(
                    start_row=min_row,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col,
                )
            elif min_row == max_row and min_col >= 2:
                ws.merge_cells(
                    start_row=min_row,
                    start_column=min_col - 1,
                    end_row=max_row,
                    end_column=max_col,
                )
            elif min_col == max_col and min_row >= 2:
                ws.merge_cells(
                    start_row=min_row - 1,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col,
                )                

    def _merged_cell_after_render(self, ws_source, ws_des, cell, new_cell):
        mergedCell = next(
            filter(
                lambda mCell: cell.coordinate in mCell,
                ws_source.merged_cells.ranges,
            ),
            None,
        )
        if mergedCell:
            diff_row = mergedCell.max_row - mergedCell.min_row
            diff_col = mergedCell.max_col - mergedCell.min_col
            ws_des.merge_cells(
                start_row=new_cell.row,
                end_row=new_cell.row + diff_row,
                start_column=new_cell.column,
                end_column=new_cell.column + diff_col,
            )

    def _context_update_child(self, context, block):
        attr = block.get("attr", None)
        child = block.get("child", None)
        if attr and child:
            try:
                context.update({child: eval(cleaning(attr), safe_globals, context)})
                return child
            except Exception as e:
                raise e
        return None

    def _check_cell_in_nested_block(self, cell, nested_block):
        if cell.row == nested_block["min_row"] and cell.column == nested_block["min_col"]:
            return 1
        elif nested_block["min_row"] <= cell.row <= nested_block["max_row"] and nested_block["min_col"] <= cell.column <= nested_block["max_col"]:
            return 0
        else:
            return -1                    

    ######################## Private Function Render Block ##########################
    def _render_block(
        self,
        ws_source,
        ws_des,
        block,
        context,
        offset_col=0,
        offset_row=0,
        max_row_by_rows=0,
        wb=None,
        left_block=None
    ):
        res = offset_col, max_row_by_rows, 0
        if block["type"] == "static":
            res = self._render_static(
                ws_source, ws_des, block, context, offset_col, offset_row, max_row_by_rows, wb, left_block)
        elif block["type"] == "forCol":
            res = self._render_col(
                ws_source, ws_des, block, context, offset_col, offset_row, max_row_by_rows, wb, left_block)
        elif block["type"] == "forRow":
            res = self._render_row(
                ws_source, ws_des, block, context, offset_col, offset_row, max_row_by_rows, wb, left_block)
        elif block["type"] == "table":
            res = self._render_table(
                ws_source, ws_des, block, context, offset_col, offset_row, max_row_by_rows, wb, left_block)        
        
        return res   

    def _render_static(self, ws_source, ws_des, block, context, offset_col=0, offset_row=0, max_row_by_rows=0, wb=None, left_block=None, from_render_col=False):
        child = self._context_update_child(context, block)

        if not left_block:
            cur_offset_col = offset_col
            new_offset_col = offset_col + (block.get("max_col", 0) - block.get("min_col", 0) + 1)
        else:
            cur_offset_col = offset_col + (block.get("min_col", 0) - left_block.get("max_col", 0) - 1)
            new_offset_col = cur_offset_col + (block.get("max_col", 0) - block.get("min_col", 0) + 1)

        new_max_row_by_rows = max(
            max_row_by_rows, block.get("max_row", 0) - block.get("min_row", 0)
        )
        nested_blocks = block.get("nested_blocks", [])
        increase_row = 0
        for row in list(
            ws_source.iter_rows(
                min_row=block.get("min_row", 0),
                max_row=block.get("max_row", 0),
                min_col=block.get("min_col", 0),
                max_col=block.get("max_col", 0),
            )
        ):
            if not row:
                continue

            for c, cell in enumerate(row):
                if isinstance(cell, MergedCell):
                    continue

                render_nested_block = False
                for nested_block in nested_blocks:
                    position_cell = self._check_cell_in_nested_block(cell, nested_block)
                    if position_cell == 0:
                        render_nested_block = True
                    elif position_cell == 1:
                        render_nested_block = True
                        if nested_block["type"] == "forRow":
                            offset_col, nested_block_new_max_row_by_rows, nested_block_increase_row = self._render_row(
                                ws_source, ws_des, nested_block, context, offset_col, offset_row,
                                max_row_by_rows, wb=wb, left_block=left_block)
                        if nested_block["type"] == "forCol":
                            offset_col, nested_block_new_max_row_by_rows, nested_block_increase_row = self._render_col(
                                ws_source, ws_des, nested_block, context, offset_col, offset_row,
                                max_row_by_rows, wb=wb, left_block=left_block)

                        # update variant
                        increase_row += nested_block_increase_row
                        new_max_row_by_rows += nested_block_new_max_row_by_rows                        
                
                if render_nested_block:
                    continue

                height = ws_source.row_dimensions[cell.row].height
                width = ws_source.column_dimensions[cell.column_letter].width
                cell_val = cell.value
                new_row = cell.row + offset_row + increase_row
                new_col = cur_offset_col + c
                if not new_row or not new_col:
                    continue

                new_cell = ws_des.cell(
                    row=new_row,
                    column=new_col,
                )
                if isinstance(new_cell, MergedCell):
                    continue

                new_cell._style = copy(cell._style)
                if check_jinja_syntax(cell_val):
                    if from_render_col:
                        new_cell.value = None
                    else:
                        render(cell_val, context, cell, new_cell, wb, ws_des)
                else:
                    new_cell.value = cell_val

                self._set_dimensions(ws_des, new_cell, height, width)
                self._merged_cell_after_render(
                    ws_source, ws_des, cell, new_cell)
        if child:
            context.pop(child)
        return new_offset_col, new_max_row_by_rows, increase_row

    def _render_col(self, ws_source, ws_des, block, context, offset_col=0, offset_row=0, max_row_by_rows=0, wb=None, left_block=None):
        new_max_row_by_rows = max_row_by_rows
        increase_row = 0
        nested_blocks = block.get("nested_blocks", [])
        nested_block = nested_blocks and nested_blocks[0] or {}
        # Get data from context
        child = self._context_update_child(context, block)
        datas = context.get(child, [])

        if not left_block:
            cur_offset_col = offset_col
            new_offset_col = offset_col + (block.get("max_col", 0) - block.get("min_col", 0) + 1) * len(datas)
        else:
            cur_offset_col = offset_col + (block.get("min_col", 0) - left_block.get("max_col", 0) - 1)
            new_offset_col = cur_offset_col + (block.get("max_col", 0) - block.get("min_col", 0) + 1) * len(datas)
        # If no data, render static content and exit
        if not datas:
            return self._render_static(ws_source, ws_des, block, context, offset_col, offset_row, max_row_by_rows, wb, left_block, from_render_col=True)

        # Cache nested block rendering results to avoid duplicate calls
        nested_render_results = {}

        for r, row in enumerate(
            ws_source.iter_rows(
                min_row=block.get("min_row", 0),
                max_row=block.get("max_row", 0),
                min_col=block.get("min_col", 0),
                max_col=block.get("max_col", 0),
            )
        ):
            for c, cell in enumerate(row):
                if isinstance(cell, MergedCell):
                    continue

                height = ws_source.row_dimensions[cell.row].height
                width = ws_source.column_dimensions[cell.column_letter].width
                cell_val = cell.value

                for i, data in enumerate(datas):
                    if child:
                        context.update({child: data})

                    render_nested_block = nested_block and (
                        nested_block["min_row"] <= row[0].row <= nested_block["max_row"])

                    # Optimize: Only render nested block once per unique offset
                    if render_nested_block and i not in nested_render_results:
                        nested_offset_col, nested_block_new_max_row_by_rows, nested_block_increase_row = self._render_row(
                            ws_source, ws_des, nested_block, context, offset_col + (i * len(row)), offset_row, max_row_by_rows, wb, left_block
                        )
                        nested_render_results[i] = (
                            nested_offset_col, nested_block_new_max_row_by_rows, nested_block_increase_row)
                        # update variant
                        if i == 0 and c == 0:
                            increase_row += nested_block_increase_row
                            new_max_row_by_rows += nested_block_new_max_row_by_rows

                    if render_nested_block:
                        continue

                    new_col = cur_offset_col + (i * len(row)) + c
                    new_row = cell.row + offset_row + increase_row

                    if not new_row or not new_col:
                        continue

                    new_cell = ws_des.cell(row=new_row, column=new_col)
                    if isinstance(new_cell, MergedCell):
                        continue

                    # Copy styles and render Jinja templates
                    new_cell._style = copy(cell._style)
                    
                    if check_jinja_syntax(cell_val):
                        render(cell.value, context, cell, new_cell, wb, ws_des)
                    else:
                        new_cell.value = cell_val

                    # Set dimensions only for the first row
                    if r == 0:
                        self._set_dimensions(ws_des, new_cell, height, width)

                    self._merged_cell_after_render(
                        ws_source, ws_des, cell, new_cell)

        # Clean up context
        if child:
            context.pop(child)

        return new_offset_col, new_max_row_by_rows, increase_row

    def _render_row(self, ws_source, ws_des, block, context, offset_col=0, offset_row=0, max_row_by_rows=0, wb=None, left_block=None):
        child = self._context_update_child(context, block)
        datas = context.get(child, [])
        if not datas:
            return offset_col, max_row_by_rows, 0
        
        if not left_block :
            new_offset_col = offset_col + (block.get("max_col", 0) - block.get("min_col", 0) + 1)
            cur_offset_col = offset_col
        else:
            cur_offset_col = offset_col + (block.get("min_col", 0) - left_block.get("max_col", 0) - 1)
            new_offset_col = cur_offset_col + (block.get("max_col", 0) - block.get("min_col", 0) + 1)
        new_max_row_by_rows = max(
            max_row_by_rows, len(datas)
        )
        increase_row = len(datas) - (block.get("max_row", 0) - block.get("min_row", 0) + 1)
        for r, row in enumerate(
            ws_source.iter_rows(
                min_row=block.get("min_row", 0),
                max_row=block.get("max_row", 0),
                min_col=block.get("min_col", 0),
                max_col=block.get("max_col", 0),
            )
        ):
            for c, cell in enumerate(row):
                if isinstance(cell, MergedCell):
                    continue
                height = ws_source.row_dimensions[cell.row].height
                cell_val = cell.value and str(cell.value).strip() or None
                for i, data in enumerate(datas):
                    new_row = cell.row + offset_row + i
                    new_col = cur_offset_col + c
                    if not new_row or not new_col:
                        continue

                    new_cell = ws_des.cell(
                        row=new_row,
                        column=new_col,
                    )
                    if isinstance(new_cell, MergedCell):
                        continue

                    new_cell._style = copy(cell._style)

                    if child:
                        context.update({child: data})

                    if check_jinja_syntax(cell_val):
                        render(cell.value, context, cell, new_cell, wb, ws_des)
                    else:
                        new_cell.value = cell_val

                    self._set_dimensions(ws_des, new_cell, height)
                    self._merged_cell_after_render(
                        ws_source, ws_des, cell, new_cell)

        if child:
            context.pop(child)
        return new_offset_col, new_max_row_by_rows, increase_row
    
    def _render_table(self, ws_source, ws_des, block, context, offset_col, offset_row, max_row_by_rows, wb, left_block):
        if block["attr"]: 
            try:
                datas = eval(cleaning(block["attr"]), safe_globals, context)
            except Exception as e:
                raise e
        else:
            datas = []

        if not datas:
            return offset_col, max_row_by_rows, 0
        
        if not left_block :
            new_offset_col = offset_col + max(len(data) for data in datas)
            cur_offset_col = offset_col
        else:
            cur_offset_col = offset_col + (block.get("min_col", 0) - left_block.get("max_col", 0) - 1)
            new_offset_col = cur_offset_col + max(len(data) for data in datas)

        row_source=block.get("min_row", 0)
        col_source=block.get("min_col", 0)

        new_max_row_by_rows = max(max_row_by_rows, len(datas))
        increase_row = len(datas) - 1

        for row, data_row in enumerate(datas):
            for col, data_cell in enumerate(data_row):
                cell = ws_source.cell(row=row_source, column=col_source)
                new_cell = ws_des.cell(row=row_source + row + offset_row, column=col + cur_offset_col)
                
                if isinstance(data_cell, Image):
                    cell_idx = f'{get_column_letter(new_cell.column)}{new_cell.row}'
                    ws_des.add_image(data_cell,cell_idx)
                    
                    if ws_des.column_dimensions[f'{get_column_letter(new_cell.column)}'].width is None or ws_des.column_dimensions[f'{get_column_letter(new_cell.column)}'].width < data_cell.width/ 8:
                        ws_des.column_dimensions[f'{get_column_letter(new_cell.column)}'].width = data_cell.width/ 8

                    if ws_des.row_dimensions[new_cell.row].height is None or ws_des.row_dimensions[new_cell.row].height < data_cell.height/ 1.3:
                        ws_des.row_dimensions[new_cell.row].height = data_cell.height/ 1.3

                else:
                    new_cell._style = copy(cell._style)
                    render(data_cell, context, cell, new_cell, wb, ws_des)

        return new_offset_col, new_max_row_by_rows, increase_row
                 