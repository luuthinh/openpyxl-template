# -*- coding:utf-8 -*-

import io
import re
import os
import logging
from ast import literal_eval
from jinja2 import Environment

from dateutil.parser import parse
from datetime import date, datetime
from PIL import Image as ImgPIL

from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.drawing.image import Image

_logger = logging.getLogger(__name__)

def is_int_string(s):
    try:
        int(s)
        return True
    except (ValueError, TypeError):
        return False

def is_date_string(s):
    try:
        parse(s)
        return True
    except (ValueError, TypeError):
        return False

def check_jinja_syntax(s):
    if not isinstance(s, str) or not s.strip():
        return False
    pattern = r"\{\{.*?\}\}"  # Matches any {{ ... }} anywhere in the string
    pattern_2 = r"\{\%.*?\%\}"  # Matches any {% ... %} anywhere in the string
    return bool(re.search(pattern, s) or re.search(pattern_2, s))

def remove_loop_syntax(s):
    if not isinstance(s, str):
        return ""    
    return re.sub(r"\{#.*?#\}", "", s)

def remove_jinja_block(s):
    if not isinstance(s, str):
        return ""
    return re.sub(r"\{\{.*?\}\}", "", s)

def get_loop_block(s):
    # Regular expression pattern to match `{# ... #}`
    if not isinstance(s, str):
        return []    
    pattern = r"\{\#(.*?)\#\}"
    return [b.replace(" ","").split(",") for b in re.findall(pattern, s)]

def cleaning(value):
    # Remove special quote
    result = re.sub(r"[“”‘’]", '"', value)
    return result


def contains_only_numbers(input_string):
    pattern = r'^[-+]?(\d+(\.\d*)?|\.\d+)$'
    return bool(re.match(pattern, str(input_string).replace(" ", "")))

def check_parentheses(s):
    return bool(re.match(r'^\s*\{.*?\}\s*$', s))

def format_value(s):
    if isinstance(s, (dict, int, float, date, datetime)):
        return s

    if not isinstance(s, str) or not s:
        return None
    try:
        if check_parentheses(s):
            return literal_eval(s)
        
        if is_int_string(s):
            return int(s)

        if ('.' in s or 'e' in s or 'E' in s) and ('-' not in s or ':' not in s):
            return float(s)              
        
        if is_date_string(s):
            return parse(s)        

        return s

    except (ValueError, TypeError):
        return s
    
def render(value, context, cell, new_cell, wb, ws_des):
    try:
        result = Environment().from_string(cleaning(str(value))).render(context)
    except Exception as e:
        pos = ""
        if cell:
            pos = f"{cell.column_letter}{cell.row}"
        raise Exception(str(e) + f" - {pos}")
    
    result = format_value(result)
    if isinstance(result, dict):
        if result.get("type") == "image":
            path = result.get("path")
            width = result.get("width") or int(ws_des.column_dimensions[new_cell.column_letter].width)
            height= result.get("height") or int(ws_des.row_dimensions[new_cell.row].height)
            img = ImgPIL.open(path)
            resized_img = img.resize((int(width * 7.5) , int(height * 1.33)))
            fp = io.BytesIO()
            resized_img.save(fp, format="png")
            ws_des.add_image(Image(fp), f'{new_cell.column_letter}{new_cell.row}')

            try:
                os.unlink(path)
            except Exception as e:
                _logger.error(f"Error when trying to remove file {path}: {e}")

        else:
            style_dict = result.get("style", None)
            font_dict = result.get("font", None)
            fill_dict = result.get("fill", None)
            border_dict = result.get("border", None)
            alignment_dict = result.get("alignment", None)
            protection_dict = result.get("protection", None)
            number_format = result.get("number_format", None)

            if style_dict and wb:
                if style_dict.get("name", "") not in wb.named_styles:
                    style = create_named_style_from_dict(**style_dict)
                    wb.add_named_style(style)
                new_cell.style = style_dict.get("name")
            if font_dict:
                new_cell.font = Font(**font_dict)
            if fill_dict:
                new_cell.fill = PatternFill(**fill_dict)
            if border_dict:
                sides = {}
                for side in ['left', 'right', 'top', 'bottom', 'diagonal', 'vertical', 'horizontal']:
                    if side in border_dict:
                        sides[side] = Side(**border_dict[side])
                new_cell.border = Border(**sides)
            if alignment_dict:
                new_cell.alignment = Alignment(**alignment_dict)
            if protection_dict:
                new_cell.protection = Protection(**alignment_dict)
            if number_format:
                new_cell.number_format = number_format
                                      
            new_cell.value = format_value(result.get("value", ""))
    else:
        new_cell.value = result

def has_parent(parent, child):
    return (
        parent['min_row'] <= child['min_row'] and parent['max_row'] >= child['max_row'] and
        parent['min_col'] <= child['min_col'] and parent['max_col'] >= child['max_col']
    )

def create_named_style_from_dict(
    name,
    font=None,
    fill=None,
    border=None,
    alignment=None,
    protection=None
):
    style = NamedStyle(name=name)

    # Font
    if font:
        style.font = Font(**font)

    # PatternFill
    if fill:
        style.fill = PatternFill(**fill)

    # Border (with Side)
    if border:
        sides = {}
        for side in ['left', 'right', 'top', 'bottom', 'diagonal', 'vertical', 'horizontal']:
            if side in border:
                sides[side] = Side(**border[side])
        style.border = Border(**sides)

    # Alignment
    if alignment:
        style.alignment = Alignment(**alignment)

    # Protection
    if protection:
        style.protection = Protection(**protection)

    return style
